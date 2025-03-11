from src.crawlers.tototuantu_crawler import TotoTuanTuCrawler
import pandas as pd
import time
import os
from src.config.settings import OUTPUT_DIR
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class TotoTuanTuExcelCrawler(TotoTuanTuCrawler):
    """Crawler chuyên dụng cho tototuantu.vn với khả năng lưu vào Excel"""
    
    def __init__(self):
        """Khởi tạo TotoTuanTuExcelCrawler"""
        super().__init__()
        self.logger.info("Khởi tạo TotoTuanTuExcelCrawler")
    
    def save_to_excel(self, products, filename):
        """
        Lưu danh sách sản phẩm vào file Excel với định dạng đẹp
        
        Args:
            products (list): Danh sách sản phẩm
            filename (str): Tên file Excel
            
        Returns:
            str: Đường dẫn đến file Excel
        """
        self.logger.info(f"Đang lưu {len(products)} sản phẩm vào Excel: {filename}")
        
        # Chuẩn bị dữ liệu cho Excel (làm phẳng các trường lồng nhau)
        flat_data = []
        
        for product in products:
            flat_product = product.copy()
            
            # Xử lý trường specifications
            if "specifications" in flat_product:
                specs = flat_product.pop("specifications", {})
                for key, value in specs.items():
                    flat_product[f"spec_{key}"] = value
            
            # Xử lý trường images
            if "images" in flat_product:
                images = flat_product.pop("images", [])
                for i, img in enumerate(images[:5]):  # Giới hạn 5 hình ảnh
                    flat_product[f"image_{i+1}"] = img
            
            # Xử lý trường all_images
            if "all_images" in flat_product:
                images = flat_product.pop("all_images", [])
                for i, img in enumerate(images[:5]):  # Giới hạn 5 hình ảnh
                    flat_product[f"image_{i+1}"] = img
            
            # Xử lý trường promotions
            if "promotions" in flat_product:
                promotions = flat_product.pop("promotions", [])
                flat_product["promotions_count"] = len(promotions)
                for i, promo in enumerate(promotions[:3]):  # Giới hạn 3 khuyến mãi
                    flat_product[f"promotion_{i+1}"] = promo
            
            flat_data.append(flat_product)
        
        # Tạo DataFrame
        df = pd.DataFrame(flat_data)
        
        # Tạo workbook và worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Thêm tiêu đề
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            # Thêm border
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # Thêm dữ liệu
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Thêm border
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
                
                # Màu nền cho hàng chẵn
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E6F0F8", end_color="E6F0F8", fill_type="solid")
        
        # Tự động điều chỉnh chiều rộng cột
        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)  # Giới hạn chiều rộng tối đa
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Đóng băng hàng đầu tiên
        ws.freeze_panes = "A2"
        
        # Lưu file
        excel_path = os.path.join(OUTPUT_DIR, filename)
        wb.save(excel_path)
        
        self.logger.info(f"Đã lưu dữ liệu vào {excel_path}")
        return excel_path
    
    def save_products_to_excel(self, products, prefix="tototuantu_products"):
        """
        Lưu danh sách sản phẩm vào file Excel
        
        Args:
            products (list): Danh sách sản phẩm
            prefix (str): Tiền tố cho tên file
            
        Returns:
            str: Đường dẫn đến file Excel
        """
        if not products:
            self.logger.warning("Không có dữ liệu sản phẩm để lưu")
            return None
        
        # Tạo tên file với timestamp
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{prefix}_{timestamp}.xlsx"
        
        # Lưu dữ liệu vào Excel
        excel_path = self.save_to_excel(products, excel_filename)
        
        return excel_path
    
    def crawl_and_save_to_excel(self, category, start_page=1, end_page=None, filter_param=None, crawl_details=True, prefix="tototuantu_products"):
        """
        Crawl sản phẩm từ một danh mục và lưu vào Excel
        
        Args:
            category (str): Tên danh mục (ví dụ: 'bon-cau', 'lavabo')
            start_page (int): Trang bắt đầu
            end_page (int, optional): Trang kết thúc. Nếu None, sẽ crawl đến trang cuối cùng.
            filter_param (str, optional): Tham số filter (nếu có)
            crawl_details (bool): Có crawl chi tiết sản phẩm hay không
            prefix (str): Tiền tố cho tên file
            
        Returns:
            tuple: (products, excel_path)
        """
        # Crawl sản phẩm
        self.logger.info(f"Bắt đầu crawl danh mục: {category}")
        products = []
        
        # Crawl danh sách sản phẩm
        products = self.crawl_tototuantu_category(
            category=category,
            start_page=start_page,
            end_page=end_page,
            filter_param=filter_param
        )
        
        # Nếu cần crawl chi tiết sản phẩm
        if crawl_details and products:
            self.logger.info(f"Bắt đầu crawl chi tiết cho {len(products)} sản phẩm")
            products_with_details = []
            
            for index, product in enumerate(products):
                if "product_url" in product and product["product_url"]:
                    self.logger.info(f"Đang crawl chi tiết sản phẩm {index + 1}/{len(products)}: {product['name']}")
                    
                    # Crawl chi tiết sản phẩm
                    product_detail = self.crawl_tototuantu_product_detail(product["product_url"])
                    
                    # Kết hợp thông tin cơ bản và chi tiết
                    if product_detail:
                        product.update(product_detail)
                
                products_with_details.append(product)
            
            products = products_with_details
        
        # Lưu dữ liệu vào Excel
        if products:
            excel_path = self.save_products_to_excel(products, prefix)
            self.logger.info(f"Đã lưu dữ liệu vào {excel_path}")
            return products, excel_path
        else:
            self.logger.warning("Không tìm thấy sản phẩm nào")
            return [], None